# LICENSE
# This software is the exclusive property of Gencovery SAS.
# The use and distribution of this software is prohibited without the prior consent of Gencovery SAS.
# About us: https://gencovery.com


from gws_core import (ConfigParams, InputSpec, InputSpecs,
                      OutputSpec, OutputSpecs, Table, Task, BoolParam, StrParam,
                      TaskInputs, TaskOutputs, task_decorator)
from gws_core.config.param.param_spec import IntParam, StrParam
from gws_core.impl.shell.shell_proxy import ShellProxy

import pandas as pd
import openpyxl
from mygene import MyGeneInfo


class NCBIGeneToENTREZIdHelper():
    CONDA_ENV_DIR_NAME = "NCBIGeneToENTREZIdCondaEnv"
    PIP_ENV_DIR_NAME = "NCBIGeneToENTREZIdPipEnv"
    CONDA_ENV_FILE_PATH = os.path.join(os.path.abspath(
        os.path.dirname(__file__)), "NCBIGeneToENTREZId_conda.yml")
    PIP_ENV_FILE_PATH = os.path.join(os.path.abspath(
        os.path.dirname(__file__)), "NCBIGeneToENTREZId_pipenv.txt")

    @classmethod
    def create_conda_proxy(cls, message_dispatcher: MessageDispatcher = None) -> CondaShellProxy:
        return CondaShellProxy(cls.CONDA_ENV_DIR_NAME, cls.CONDA_ENV_FILE_PATH, message_dispatcher=message_dispatcher)

    @classmethod
    def create_pip_proxy(cls, message_dispatcher: MessageDispatcher = None) -> PipShellProxy:
        return PipShellProxy(cls.PIP_ENV_DIR_NAME, cls.PIP_ENV_FILE_PATH, message_dispatcher=message_dispatcher)



@task_decorator("NCBIGeneToENTREZId", human_name='NCBI Gene to ENTREZ ID', short_description="ADD a new column gene_id corresponding to the gene ENTREZ gene ID")
class NCBIGeneToENTREZId(task):

    input_specs = InputSpecs({'input_table': InputSpec(Table, human_name="input_table")})
    output_specs = OutputSpecs({'output_table': OutputSpec(Table, human_name="output_table")})

    config_specs = {}

    def run(self, params: ConfigParams, inputs: TaskInputs) -> TaskOutputs:
        dataframe = pd.DataFrame(inputs['input_table'].get_data())

        # Load the input file
        input_file = source_paths[0]

        def convert_gene_names_to_ids(gene_names):
            mg = MyGeneInfo()
            results = mg.querymany(gene_names, scopes='symbol', fields='entrezgene', species='human', returnall=True)

            gene_ids = {}
            for result in results['out']:
                if 'entrezgene' in result:
                    gene_name = result['query']
                    gene_id = result['entrezgene']
                    gene_ids[gene_name] = gene_id

            return gene_ids

        def add_gene_ids_to_xlsx(input_file, output_file):

            sheet = dataframe

            gene_names = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1)]
            #sheet with all the gene names from column 1 from 2nd row to last one
            gene_ids = convert_gene_names_to_ids(gene_names)

            sheet.insert_cols(2)  # Insert a new column for gene IDs
            sheet.cell(row=1, column=2, value='gene_id')  # Set the column header as 'gene_id'

            for i, gene_name in enumerate(gene_names, start=2):
                gene_id = gene_ids.get(gene_name)

                sheet.cell(row=i, column=2, value=gene_id)

            wb.save(output_file)

        # Example usage
        output_file = 'converted_gene_ids.xlsx'
        add_gene_ids_to_xlsx(input_file, output_file)

        # Convert the XLSX file to a CSV file while preserving the column name
        df = pd.read_excel(output_file)
        df.to_csv('converted_gene_ids.csv', index=False, header=True)

        target_paths = ['converted_gene_ids.csv']
